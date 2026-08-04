"""Download, inspect, redact, and format Excel workbooks for the UI."""

from __future__ import annotations

import hashlib
import json
import re
from datetime import date, datetime, time
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles.colors import COLOR_INDEX
from openpyxl.styles.numbers import is_date_format
from openpyxl.utils import get_column_letter

from config import Settings


class WorkbookError(RuntimeError):
    pass


# Browser-based downloading is handled by downloader.py
# WorkbookProcessor only processes local workbook files.

class WorkbookProcessor:
    MAX_SAMPLE_ROWS = 5
    MAX_SAMPLE_COLUMNS = 20
    MAX_PREVIEW_ROWS = 5000
    MAX_PREVIEW_COLUMNS = 200
    PII_PATTERNS = (
        ("ssn_hyphen", re.compile(r"(?<!\d)\d{3}-\d{2}-\d{4}(?!\d)")),
        ("ssn_dot", re.compile(r"(?<!\d)\d{3}\.\d{2}\.\d{4}(?!\d)")),
        ("ssn_space", re.compile(r"(?<!\d)\d{3}\s\d{2}\s\d{4}(?!\d)")),
        ("card_plain", re.compile(r"(?<!\d)\d{16}(?!\d)")),
        ("card_hyphen", re.compile(r"(?<!\d)\d{4}-\d{4}-\d{4}-\d{4}(?!\d)")),
        ("card_dot", re.compile(r"(?<!\d)\d{4}\.\d{4}\.\d{4}\.\d{4}(?!\d)")),
        ("card_space", re.compile(r"(?<!\d)\d{4}\s\d{4}\s\d{4}\s\d{4}(?!\d)")),
    )
    RULES_DIR = Path(__file__).resolve().parents[1] / "data" / "sheet_rules"
    CONTEXT_DIR = Path(__file__).resolve().parents[1] / "data" / "sheet_contexts"
    MAPPING_PATH = RULES_DIR / "mappings.json"
    _template_cache: dict[str, dict] | None = None
    _mapping_cache: dict | None = None

    def build_payload(self, workbook_path: Path, workfront_content: dict | str | None = None) -> dict:
        if isinstance(workfront_content, str):
            workfront_content = {"OFFER_DESCRIPTION": workfront_content}
        workfront_content = {
            column: self._json_safe_value(value)
            for column, value in (workfront_content or {}).items()
        }
        workfront_text = "\n".join(
            f"{column}: {'N/A' if value is None or str(value).strip() == '' else value}"
            for column, value in workfront_content.items()
        )
        try:
            workbook = load_workbook(workbook_path, read_only=False, data_only=True)
        except Exception as error:
            raise WorkbookError("The downloaded file is not a readable Excel workbook.") from error

        pii_count = 0
        sheet_data: dict[str, dict] = {}
        try:
            for sheet in workbook.worksheets:
                rendered, found = self._sheet_sample(sheet)
                pii_count += found
                preview_rows = self._preview_rows(sheet)
                dimensions = f"{sheet.max_row} rows × {sheet.max_column} columns"
                template = self.resolve_worksheet_template(sheet.title)
                worksheet_type = template["worksheet_type"]
                sheet_data[sheet.title] = {
                    "worksheet_type": worksheet_type,
                    "validation_template": template["template_name"],
                    "validation_summary": template["categories"][:5],
                    "offer_description": workfront_text,
                    "workfront_content": workfront_content,
                    "workbook_context": self._sheet_context(
                        workbook_path.name, sheet.title, workbook.sheetnames
                    ),
                    "sheet_prompt": self._sheet_instructions(sheet.title, template),
                    "sheet_sample": rendered,
                    "dimensions": dimensions,
                    "generated_prompt": self._generated_prompt(
                        workfront_text, workbook_path.name, sheet.title,
                        template, dimensions, rendered
                    ),
                    "preview_rows": preview_rows["rows"],
                    "preview_layout": preview_rows["layout"],
                    "preview_truncated": (
                        sheet.max_row > self.MAX_PREVIEW_ROWS
                        or sheet.max_column > self.MAX_PREVIEW_COLUMNS
                    ),
                }
        finally:
            workbook.close()

        return {
            "workbook_name": workbook_path.name,
            "workbook_type": workbook_path.suffix.removeprefix(".").upper() or "XLSX",
            "sheet_count": len(workbook.sheetnames),
            "sheets": workbook.sheetnames,
            "sheet_data": sheet_data,
            "workfront_content": workfront_content,
            "has_pii": pii_count > 0,
            "pii_count": pii_count,
            "pii_status": f"{pii_count} detected for masking" if pii_count else "None detected"
        }

    @staticmethod
    def _json_safe_value(value):
        if value is None or isinstance(value, (str, int, float, bool)):
            return value
        if hasattr(value, "isoformat"):
            return value.isoformat()
        if hasattr(value, "read"):
            return str(value.read())
        return str(value)

    def _sheet_sample(self, sheet) -> tuple[str, int]:
        rows: list[str] = []
        pii_count = 0
        for row in sheet.iter_rows(max_row=self.MAX_SAMPLE_ROWS, max_col=self.MAX_SAMPLE_COLUMNS, values_only=True):
            values = []
            for value in row:
                rendered_value = "" if value is None else str(value)
                pii_count += self._count_pii(rendered_value)
                values.append(rendered_value)
            rows.append(" | ".join(values).rstrip(" |"))
        return "\n".join(rows).strip() or "(The sheet is empty.)", pii_count

    def _preview_rows(self, sheet) -> dict:
        max_row = min(sheet.max_row, self.MAX_PREVIEW_ROWS)
        max_column = min(sheet.max_column, self.MAX_PREVIEW_COLUMNS)
        rows = []
        styles = {}
        for row in sheet.iter_rows(max_row=max_row, max_col=max_column):
            rendered_row = []
            for cell in row:
                style_id = str(cell.style_id)
                if style_id not in styles:
                    styles[style_id] = self._cell_style(cell)
                rendered_row.append({
                    "value": self._display_value(cell.value, cell.number_format),
                    "style_id": style_id,
                })
            rows.append(rendered_row)

        merged_ranges = []
        for merged in sheet.merged_cells.ranges:
            if merged.min_row > max_row or merged.min_col > max_column:
                continue
            merged_ranges.append({
                "min_row": merged.min_row,
                "max_row": min(merged.max_row, max_row),
                "min_col": merged.min_col,
                "max_col": min(merged.max_col, max_column),
            })

        return {
            "rows": rows,
            "layout": {
                "column_widths": [
                    sheet.column_dimensions[get_column_letter(column)].width
                    for column in range(1, max_column + 1)
                ],
                "row_heights": [
                    sheet.row_dimensions[row].height
                    for row in range(1, max_row + 1)
                ],
                "merged_ranges": merged_ranges,
                "show_gridlines": sheet.sheet_view.showGridLines is not False,
                "styles": styles,
            },
        }

    @staticmethod
    def _display_value(value, number_format: str) -> str:
        if value is None:
            return ""
        if isinstance(value, (datetime, date, time)) or (
            isinstance(value, (int, float)) and is_date_format(number_format)
        ):
            if isinstance(value, datetime):
                return value.strftime("%m/%d/%Y %H:%M") if "h" in number_format.lower() else value.strftime("%m/%d/%Y")
            if isinstance(value, date):
                return value.strftime("%m/%d/%Y")
            if isinstance(value, time):
                return value.strftime("%H:%M:%S")
        if isinstance(value, bool):
            return "TRUE" if value else "FALSE"
        if isinstance(value, (int, float)):
            format_code = (number_format or "General").split(";")[0]
            if format_code.lower() == "general":
                return str(value)
            if "%" in format_code:
                decimals = len(format_code.split(".", 1)[1].split("%", 1)[0]) if "." in format_code else 0
                return f"{value * 100:,.{decimals}f}%"
            decimals = len(re.search(r"\.([0#]+)", format_code).group(1)) if re.search(r"\.([0#]+)", format_code) else 0
            use_grouping = "," in format_code
            rendered = f"{value:,.{decimals}f}" if use_grouping else f"{value:.{decimals}f}"
            currency = next((symbol for symbol in ("$", "€", "£", "¥", "₹") if symbol in format_code), "")
            return f"{currency}{rendered}"
        return str(value)

    @classmethod
    def _cell_style(cls, cell) -> dict:
        style = {
            "number_format": cell.number_format,
            "font": {
                "name": cell.font.name,
                "size": cell.font.sz,
                "bold": bool(cell.font.bold),
                "italic": bool(cell.font.italic),
                "underline": bool(cell.font.underline),
                "color": cls._color_to_css(cell.font.color),
            },
            "fill": cls._color_to_css(cell.fill.fgColor) if cell.fill.fill_type == "solid" else None,
            "alignment": {
                "horizontal": cell.alignment.horizontal,
                "vertical": cell.alignment.vertical,
                "wrap_text": bool(cell.alignment.wrap_text),
                "text_rotation": cell.alignment.text_rotation,
            },
            "borders": {},
        }
        for edge in ("top", "right", "bottom", "left"):
            side = getattr(cell.border, edge)
            if side and side.style:
                style["borders"][edge] = {
                    "style": side.style,
                    "color": cls._color_to_css(side.color) or "#000000",
                }
        return style

    @staticmethod
    def _color_to_css(color) -> str | None:
        if not color:
            return None
        if color.type == "rgb" and color.rgb:
            return f"#{color.rgb[-6:]}"
        if color.type == "indexed" and color.indexed is not None:
            rgb = COLOR_INDEX[color.indexed] if color.indexed < len(COLOR_INDEX) else None
            return f"#{rgb[-6:]}" if rgb else None
        if color.type == "theme" and color.theme is not None:
            theme_colors = (
                "#FFFFFF", "#000000", "#E7E6E6", "#44546A",
                "#4472C4", "#ED7D31", "#A5A5A5", "#FFC000",
                "#5B9BD5", "#70AD47", "#0563C1", "#954F72",
            )
            return theme_colors[color.theme] if color.theme < len(theme_colors) else None
        return None

    def _count_pii(self, text: str) -> int:
        return sum(
            len(pattern.findall(text))
            for _, pattern in self.PII_PATTERNS
        )

    def _mask_pii(self, text: str) -> tuple[str, int]:
        matches = 0
        for _, pattern in self.PII_PATTERNS:
            text, count = pattern.subn("[REDACTED]", text)
            matches += count
        return text, matches

    @staticmethod
    def _workbook_context(name: str, sheets: list[str]) -> str:
        return f"Workbook: {name}\nAvailable sheets: {', '.join(sheets)}\nCredit Card and SSN values will be masked when the prompt is generated."

    @staticmethod
    def normalize_sheet_name(name: str) -> str:
        """Normalize a worksheet title before applying keyword rules."""
        normalized = str(name).lower().replace("_", " ").replace("-", " ")
        normalized = re.sub(r"[^a-z0-9\s]", " ", normalized)
        return re.sub(r"\s+", " ", normalized).strip()

    @classmethod
    def _load_templates(cls) -> dict[str, dict]:
        if cls._template_cache is not None:
            return cls._template_cache
        templates = {}
        for path in sorted(cls.RULES_DIR.glob("*.txt")):
            metadata, body = cls._parse_template(path.read_text(encoding="utf-8"))
            worksheet_type = metadata.get("worksheet_type", path.stem)
            templates[path.stem] = {
                "key": path.stem,
                "worksheet_type": worksheet_type,
                "template_name": metadata.get("template", f"{worksheet_type} Validation"),
                "categories": [
                    item.strip()
                    for item in metadata.get("categories", "").split("|")
                    if item.strip()
                ][:5],
                "content": body.strip(),
            }
        cls._template_cache = templates
        return templates

    @staticmethod
    def _parse_template(text: str) -> tuple[dict[str, str], str]:
        metadata = {}
        body_lines = []
        reading_metadata = True
        for line in text.splitlines():
            if reading_metadata and line.startswith("@") and ":" in line:
                key, value = line[1:].split(":", 1)
                metadata[key.strip().lower()] = value.strip()
                continue
            if reading_metadata and not line.strip():
                continue
            reading_metadata = False
            body_lines.append(line)
        return metadata, "\n".join(body_lines)

    @classmethod
    def _load_mapping(cls) -> dict:
        if cls._mapping_cache is None:
            cls._mapping_cache = json.loads(cls.MAPPING_PATH.read_text(encoding="utf-8"))
        return cls._mapping_cache

    @classmethod
    def resolve_worksheet_template(cls, sheet_name: str) -> dict:
        templates = cls._load_templates()
        mapping = cls._load_mapping()
        sheet_words = set(cls.normalize_sheet_name(sheet_name).split())
        matches = []
        for template_key, template in templates.items():
            keywords = mapping.get("templates", {}).get(template_key, [])
            keywords = [*keywords, template["worksheet_type"], template_key]
            for keyword in keywords:
                keyword_words = set(cls.normalize_sheet_name(keyword).split())
                if keyword_words and keyword_words.issubset(sheet_words):
                    matches.append((len(keyword_words), len(keyword), template_key))
        if matches:
            return templates[max(matches)[2]]
        fallback = mapping.get("fallback", "Generic Validation")
        if fallback not in templates:
            raise WorkbookError(f"Generic worksheet template '{fallback}.txt' was not found.")
        return templates[fallback]

    @classmethod
    def detect_worksheet_type(cls, sheet_name: str) -> str:
        return cls.resolve_worksheet_template(sheet_name)["worksheet_type"]

    @classmethod
    def _sheet_instructions(cls, sheet_name: str, template: dict) -> str:
        return template["content"].format(
            sheet_name=sheet_name,
            worksheet_type=template["worksheet_type"],
        )

    @classmethod
    def _sheet_context(cls, workbook_name: str, sheet_name: str, sheets: list[str]) -> str:
        worksheet_type = cls.detect_worksheet_type(sheet_name)
        configured = cls._load_sheet_text(cls.CONTEXT_DIR, worksheet_type)
        if configured:
            return configured
        return (
            f"Workbook: {workbook_name}\n"
            f"Active worksheet: {sheet_name}\n"
            f"Available sheets: {', '.join(sheets)}\n"
            "Credit Card and SSN values will be masked when the prompt is generated."
        )

    @staticmethod
    def _load_sheet_text(directory: Path, sheet_name: str) -> str:
        filename = re.sub(r"[^a-z0-9]+", "_", sheet_name.lower()).strip("_") + ".txt"
        path = directory / filename
        if not path.is_file():
            return ""
        return path.read_text(encoding="utf-8").strip()

    @classmethod
    def _generated_prompt(
        cls, offer: str, name: str, sheet: str, template: dict,
        dimensions: str, sample: str
    ) -> str:
        worksheet_type = template["worksheet_type"]
        instructions = cls._sheet_instructions(sheet, template)
        prompt = (
            "# Workbook Validation Prompt\n\n"
            f"## Offer Description\n{offer or '(Not available)'}\n\n"
            f"## Workbook Context\nWorkbook: {name}\nSheet: {sheet} ({dimensions})\n"
            f"Worksheet type: {worksheet_type}\n\n"
            f"## Validation Instructions\n{instructions}\n\n"
            f"## Masked Sheet Sample\n{sample}\n\n"
            "## Expected Output\nList validation issues with row/column references, severity, and suggested remediation."
        )
        return cls()._mask_pii(prompt)[0]
