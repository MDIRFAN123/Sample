from pathlib import Path

from openpyxl import Workbook

from services.workbook_service import WorkbookProcessor


def test_detects_normalized_worksheet_types():
    processor = WorkbookProcessor()
    examples = {
        "Waterfall": "Waterfall",
        "Waterfall Sheet": "Waterfall",
        "Waterfall_v2": "Waterfall",
        "Campaign Waterfall": "Waterfall",
        "Offer Description": "Offer Description",
        "Offer Desc": "Offer Description",
        "Offer Details": "Offer Description",
        "Landing Page": "Landing Page",
        "Landing Pages": "Landing Page",
        "Terms": "Terms",
        "Terms & Conditions": "Terms",
        "Email": "Email",
        "Email Content": "Email",
        "Unexpected Data": "Generic Workbook",
    }

    for sheet_name, expected_type in examples.items():
        assert processor.detect_worksheet_type(sheet_name) == expected_type


def test_build_payload_supports_all_sheets_and_renders_prompt(tmp_path: Path):
    workbook_path = tmp_path / "dynamic.xlsx"
    workbook = Workbook()
    workbook.active.title = "Campaign Waterfall"
    workbook.create_sheet("Landing Pages")
    workbook.create_sheet("Custom Inputs")
    workbook.save(workbook_path)

    payload = WorkbookProcessor().build_payload(workbook_path)

    assert payload["sheets"] == [
        "Campaign Waterfall",
        "Landing Pages",
        "Custom Inputs",
    ]
    assert payload["sheet_count"] == 3
    assert payload["sheet_data"]["Campaign Waterfall"]["worksheet_type"] == "Waterfall"
    assert payload["sheet_data"]["Landing Pages"]["worksheet_type"] == "Landing Page"
    generic = payload["sheet_data"]["Custom Inputs"]
    assert generic["worksheet_type"] == "Generic Workbook"
    assert "Worksheet: Custom Inputs" in generic["sheet_prompt"]
    assert "Detected worksheet type: Generic Workbook" in generic["sheet_prompt"]
    assert "{validation_rules}" not in generic["sheet_prompt"]
